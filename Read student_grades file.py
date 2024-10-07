# Read and retrieve students grades in an excel file

from openpyxl import load_workbook
import openpyxl

workbook  = load_workbook("students_grades.xlsx")

sheet = workbook.active

print("Students with grades above 85: ")

for row in sheet.iter_rows(min_row=2, values_only=True):
    name, grade = row
    if grade > 85:
        print(f"{name}: {grade}")