# multiply all numeric values by 10 from one file and save the updated vaues in a new file

from openpyxl import Workbook, load_workbook
import openpyxl

def create_numerics():
    workbook = Workbook()
    worksheet = workbook.active
    
    list_values = ["Value"]
    num_of_values = int(input("Enter the maximum number of values to input: "))
    
    for i in range(num_of_values):
        value = float(input(f"Enter value {i+1}: "))
        list_values.append([value])
        
    for value in list_values:
        worksheet.append(value)
        
    workbook.save("numerics.xlsx")
    
    
def multiply_values():
    workbook = load_workbook("numerics.xlsx")
    worksheet = workbook.active
    
    worksheet['B1'] = "Multiplication_values"
    
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        for value in row:
            multiplication_value = value * 10
            worksheet.append(multiplication_value)
            
    workbook.save("numerics.xlsx")
    
    
create_numerics()
multiply_values()