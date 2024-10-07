# Current time spread sheet

import time
from datetime import datetime
from openpyxl import Workbook, load_workbook

def create_excel_with_datetime():
    workbook = Workbook()
    worksheet = workbook.active
    
    worksheet['A1'] = "Data"
    worksheet['B1'] = "Timestamp"
    worksheet['A2'] = "Example"
    worksheet['B2'] = datetime.now().strftime("%Y-%M-%D %H:%M:%S")
    
    workbook.save("date_time_record.xlsx")
    print("Excel file created with current date and time!")
    
def update_excel_with_last_modified():
    workbook = load_workbook("date_time_record.xlsx")
    worksheet = workbook.active
    
    worksheet['C1'] = "Last Modfied"
    worksheet['C2'] = datetime.now().strftime("%Y-%M-%D %H:%M:%S")
    
    workbook.save("date_time_record.xlsx")
    print("Excel file updated with 'Last Modified' date!")
    

create_excel_with_datetime()
update_excel_with_last_modified()


