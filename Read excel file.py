# read or retrieve data from an excel file

from openpyxl import Workbook, load_workbook

wb = load_workbook("product_inventory.xlsx")

sh = wb.active

for row in sh.iter_rows(values_only=True):
    print(row)